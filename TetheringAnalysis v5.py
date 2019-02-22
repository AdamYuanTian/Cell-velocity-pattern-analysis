#This code is used to identify the cells that go through tethering from all the tracked cells.

#Version 3 README (Old)
#Currently, the defination of tethering here is:
#	1. For a cell, its instantaneous velocities are lower than a specific cutoff velocity for a certain length of frames.
#		Eg. The instantaneous velocity of a cell is lower than 10 um/s for 3 continuous frames.
#Update needed: 1. For a tethering cell, if the starting velocity is smaller than a specific value, it should not be included in the final list.
#                  Since the cell is not moving at all.
#               2. Instantaneous velocity is calculated using ((x2-x1)^2+(y2-y1)^2)^0.5, so the cell area fluctuation caused by cell overlap
#                  will influence cell location (center of area), thus influence instantaneous velocity. The fluctuation should be excluded.

#Version 4 README
#Currently, the defination of tethering here is:
#	1. (Max&Min) For the instantaneous velocities of a cell, if the difference between Min and Max is equal or smaller than a cutoff value (negative), 
#	   and the Min is after Max, the cell is considered tethering. If the Min is not after Max, 2nd round assessment is needed.
#	2. (Dynamic Baseline) A dynamic velocity baseline is used, and is always replaced by a higher value when going through all the instantaneous velocities.
#	   When the difference between Current Velocity and the Baseline reaches the cutoff (a negative value), the cell is considered tethering.
#	Note: The code will go through all information, but currently we are only looking at Rolled Cells at Shear Rate 20 s-1

#Version 5 README
#A summary file is added with Velocity Fluctuation Analysis included.
#The summary format is good for Minitab analysis.
#More criteria is needed for defining tethering cells.

########### Place the py file in the folder with excel files need to be analyzed.

import os
import openpyxl
import xlwt
import numpy as np 
import pandas as pd 
from pandas import ExcelWriter
from pandas import ExcelFile
from scipy.stats import kurtosis, skew 

########### Custom parameters for analysis
#continuousVelocityLength = int(input('Please type in the cutoff velocity difference for tethering: '))
#velocityDiffCutoff = int(input('Please type in the cutoff velocity difference for cell tethering: '))
velocityDiffCutoff = -100 #-100 um/s is used currently for shear rate 20 s-1

########### Get the list of directory of files needs to be analyzed
sourceDir = os.getcwd() #get current working directory
items = os.listdir(".") #get all the file names under this folder
folderList = []
pathList = []
for names in items:
	if names.endswith(".xlsx"): #get all the xlsx files under this folder
		folderList.append(names) #add all the xlsx files to the folderList
		pathList.append(sourceDir + '\\' + names) #add all the directory of xlsx files to the pathList

########## Check each file in the folder
sheetInformation_allFiles = [] #Store calculation results for all files
allCellName_allFiles = []
sheetNames_allFiles = []
aveVelocity_allFiles = []
stdVelocity_allFiles = []
aveLogVelocity_allFiles = []
stdLogVelocity_allFiles = []
aveLogVelocityDiff_allFiles = []
stdLogVelocityDiff_allFiles = []
skewLogVelocityDiff_allFiles = []
kurtLogVelocityDiff_allFiles = []


for fileName in folderList:
	print ('File: ' + fileName + ' is being processed...')
	wb = openpyxl.load_workbook(fileName) #load the file

########## Check each sheet in the file
	tetheringCellName_allSheets=[]
	tetheringCellVelocity_allSheets=[]
	
	sheetInformation = [] #Store calculation results for each file
	allCellName_allSheets = []
	aveVelocity_allSheets = []
	stdVelocity_allSheets = []
	aveLogVelocity_allSheets = []
	stdLogVelocity_allSheets = []
	aveLogVelocityDiff_allSheets = []
	stdLogVelocityDiff_allSheets = []
	skewLogVelocityDiff_allSheets = []
	kurtLogVelocityDiff_allSheets = []


	sheetNames = wb.sheetnames #load all the sheet names and save to sheetNames
	for sheet in sheetNames: #load each sheet in the file
		if sheet=='Sheet1': #'Sheet1' is blank
			continue
		print ('Sheet: ' + sheet + ' is being processed...')
		df = pd.read_excel(fileName, sheet_name=sheet) #store all data in current sheet into dataframe

########## Store each cell velocities from the sheet into a 2D list of list
		listVelocity = df['Velocity (um/s)'] #store the whole velocity column into a list
		listLabel = df['Label'] #store the whole label column into a list
		listCellStatus = df['Cell Status'] #store the cell status into a list
		cellVelocity_temp = [] #for storing current cell velocity temporarily
		allCellVelocity = [] #for storing all cell velocity in 2D list
		allCellName = [] #for storing all cell names
		
		for i in range(len(listVelocity)): #go through each rows
			if listVelocity[i] == 0 and pd.notnull(listCellStatus[i]): #sometines the velocity can be zero in the middle, so make sure cell status is not NaN
				allCellVelocity.append(cellVelocity_temp) #add one cell velocity to 2D list for all cells
				cellVelocity_temp=[] #reset
				allCellName.append(listLabel[i]) #add current cell name to list
				continue
			cellVelocity_temp.append(listVelocity[i])
		
		allCellVelocity.append(cellVelocity_temp) #add the last cell's velocities to the 2D list
		allCellVelocity.pop(0) #clear the first empty element
		
########### Read information from current sheet for data output later
		sheetInformation_temp = []
		sheetInformation_temp = allCellName[0].split()
		info = []
		for k in range (0, 6): #get the first 6 information name
			info.append(sheetInformation_temp[k])
		sheetInformation.append(info) #cellType material concentration expDate trial shearRate

########### Analyze each cell velocities for finding tethering
		aveVelocity = [] #store calculation result for current sheet
		stdVelocity = []
		aveLogVelocity = []
		stdLogVelocity = []
		aveLogVelocityDiff = []
		stdLogVelocityDiff = []
		skewLogVelocityDiff = []
		kurtLogVelocityDiff = []

		tetheringCellVelocity = [] #store the tethering cell's velocities in current sheet
		tetheringCellName = []
		for j in range(len(allCellVelocity)): #check each cell in the 2D list
			############# Changes needed for tethering criteria
			#for n in range(len(allCellVelocity[j])-continuousVelocityLength): #check each velocity for the cell
			#	count = 0
			#	if allCellVelocity[j][n] < velocityCutoff:
			#		for m in range(continuousVelocityLength): #check the value of continuous number with the length of continuousVelocityLength
			#			if allCellVelocity[j][n+m] < velocityCutoff:
			#				count = count + 1
			#	if count == continuousVelocityLength: #compare the length 
			#		tetheringCellVelocity.append(allCellVelocity[j]) #store the velocity of tethering cells
			#		tetheringCellName.append(allCellName[j]) #store the name of tethering cells
			#		break
		
			######Criteria 1, compare the Max and Min instantaneous velocity of a cell, and compare their order
			maxInsVelocity = max(allCellVelocity[j]) #Get the max velocity for a cell
			maxInsVelocityLoc = allCellVelocity[j].index(max(allCellVelocity[j])) #Get the location of the 1st max velocity, assuming there is only one max
			
			minInsVelocity = min(allCellVelocity[j]) #Get the min velocity for a cell              
			minInsVelocityLoc = [] #Store all the locations of min velocity
			for counter, value in enumerate(allCellVelocity[j]): #Get all the locations of min velocity, there maybe multiple zero velocities for a cell
				if value == minInsVelocity:
					minInsVelocityLoc.append(counter)
			
			if minInsVelocity - maxInsVelocity <= velocityDiffCutoff and maxInsVelocityLoc < minInsVelocityLoc[-1]: #Apply Criteria 1
				tetheringCellVelocity.append(allCellVelocity[j]) #Store the velocity of tethering cells
				tetheringCellName.append(allCellName[j]) #Store the name of tethering cells
				continue #Skip Criteria 2 if Critera 1 is fulfilled
				
			######Criteria 2, use a Dynamic Baseline to assess the drop in instantaneous velocity
			baseline = allCellVelocity[j][0] #Initialize the baseline with the 1st instantaneous velocity
			
			for n in range(len(allCellVelocity[j])): #Go through each instantaneous velocity
				velocityDiff = allCellVelocity[j][n]-baseline
				if velocityDiff > 0: #Replace the baseline with a higher value
					baseline = allCellVelocity[j][n] #Dynamic Baseline
					continue #Proceed to next loop
				if velocityDiff <= velocityDiffCutoff: #Meet the 2nd criteria
					tetheringCellVelocity.append(allCellVelocity[j]) #Store the velocity of tethering cells
					tetheringCellName.append(allCellName[j]) #Store the name of tethering cells
					break #Terminate the loop
					
			######Calculation for assessing velocity fluctuation
			aveVelocity_temp = np.mean(allCellVelocity[j]) 
			aveVelocity.append(aveVelocity_temp) #store result for each cell

			stdVelocity_temp = np.std(allCellVelocity[j])
			stdVelocity.append(stdVelocity_temp)

			logVelocity = [] #transfer velocity to natural log
			logVelocityDiff = [] #calculate log idfference of velocity

			for m in range(len(allCellVelocity[j])):
				if allCellVelocity[j][m] == 0: #replace zero velocity with one
					logVelocity.append(np.log(1))
					continue
				logVelocity.append(np.log(allCellVelocity[j][m]))

			aveLogVelocity_temp = np.mean(logVelocity)
			aveLogVelocity.append(aveLogVelocity_temp)

			stdLogVelocity_temp = np.std(logVelocity)
			stdLogVelocity.append(stdLogVelocity_temp)

			for o in range(len(logVelocity)-1): #log difference calculation
				logVelocityDiff.append(logVelocity[o+1]-logVelocity[o])

			aveLogVelocityDiff_temp = np.mean(logVelocityDiff)
			aveLogVelocityDiff.append(aveLogVelocityDiff_temp)

			stdLogVelocityDiff_temp = np.std(logVelocityDiff)
			stdLogVelocityDiff.append(stdLogVelocityDiff_temp)

			skewLogVelocityDiff_temp = skew(logVelocityDiff)
			skewLogVelocityDiff.append(skewLogVelocityDiff_temp)

			kurtLogVelocityDiff_temp = kurtosis(logVelocityDiff)
			kurtLogVelocityDiff.append(kurtLogVelocityDiff_temp)

			##################Inlcude tethering information##########

		tetheringCellName_allSheets.append(tetheringCellName)
		tetheringCellVelocity_allSheets.append(tetheringCellVelocity)

		aveVelocity_allSheets.append(aveVelocity) #store result for each sheet
		stdVelocity_allSheets.append(stdVelocity)
		aveLogVelocity_allSheets.append(aveLogVelocity)
		stdLogVelocity_allSheets.append(stdLogVelocity)
		aveLogVelocityDiff_allSheets.append(aveLogVelocityDiff)
		stdLogVelocityDiff_allSheets.append(stdLogVelocityDiff)
		skewLogVelocityDiff_allSheets.append(skewLogVelocityDiff)
		kurtLogVelocityDiff_allSheets.append(kurtLogVelocityDiff)
		allCellName_allSheets.append(allCellName)

	aveVelocity_allFiles.append(aveVelocity_allSheets) #store result for each file
	stdVelocity_allFiles.append(stdVelocity_allSheets)
	aveLogVelocity_allFiles.append(aveLogVelocity_allSheets)
	stdLogVelocity_allFiles.append(stdLogVelocity_allSheets)
	aveLogVelocityDiff_allFiles.append(aveLogVelocityDiff_allSheets)
	stdLogVelocityDiff_allFiles.append(stdLogVelocityDiff_allSheets)
	skewLogVelocityDiff_allFiles.append(skewLogVelocityDiff_allSheets)
	kurtLogVelocityDiff_allFiles.append(kurtLogVelocityDiff_allSheets)
	allCellName_allFiles.append(allCellName_allSheets)
	sheetNames_allFiles.append(sheetNames)
	sheetInformation_allFiles.append(sheetInformation)

	#####write the data into excel and move onto next file
	wkbk = xlwt.Workbook()
	for a in range(len(sheetNames)-1):  #'Sheet1' is not included because it is blank
		ws = wkbk.add_sheet(sheetNames[a+1]) #add sheet to file
		for b in range(len(tetheringCellName_allSheets[a])):
			ws.write(0,b,tetheringCellName_allSheets[a][b]) #add cell name to the first row
			for c in range(len(tetheringCellVelocity_allSheets[a][b])):
				ws.write(c+1,b,tetheringCellVelocity_allSheets[a][b][c]) #add value into each cell, a is the sheet, b is the cell, c is the velocity
	wkbk.save(fileName[:-5]+'_Tethering Analyzed'+'.xls')


	print ('\n') #space between files


#####Write analyzed data into a summary excel file
wkbk_sum = xlwt.Workbook()
ws1 = wkbk_sum.add_sheet('tetheringAnalysisSummary')
title = ['CellType', 'Material', 'Concentration', 'ExpDate', 'Trial', 'ShearRate', 'CellName', 'SheetNames', 'Rolling', 'AveVelocity (um/s)', 'STD(v)', 'AveLog(v)', 'STD(log(v))', 'AveLog(v)Diff', 'STD(Log(v)Diff)', 'Skew(Log(v)Diff)', 'Kurt(Log(v)Diff)']
for d in range(len(title)):  #write title on the first row
	ws1.write(0, d, title[d])

row = 1
col = 0

for e in range(len(aveVelocity_allFiles)): #loop for each file (material)
	for f in range(len(aveVelocity_allFiles[e])): #loop for each sheet (shear rate, rolled or nonrolled)
		for g in range(len(aveVelocity_allFiles[e][f])):
			ws1.write(row, col, sheetInformation_allFiles[e][f][col])
			ws1.write(row, col+1, sheetInformation_allFiles[e][f][col+1])
			ws1.write(row, col+2, sheetInformation_allFiles[e][f][col+2])
			ws1.write(row, col+3, sheetInformation_allFiles[e][f][col+3])
			ws1.write(row, col+4, sheetInformation_allFiles[e][f][col+4])
			ws1.write(row, col+5, sheetInformation_allFiles[e][f][col+5])
			ws1.write(row, col+6, allCellName_allFiles[e][f][g])
			ws1.write(row, col+7, sheetNames_allFiles[e][f+1])
			if 'Non' in sheetNames_allFiles[e][f+1]:
				ws1.write(row, col+8, 'NonRolling')
			else:
				ws1.write(row, col+8, 'Rolling')
			ws1.write(row, col+9, aveVelocity_allFiles[e][f][g])
			ws1.write(row, col+10, stdVelocity_allFiles[e][f][g])
			ws1.write(row, col+11, aveLogVelocity_allFiles[e][f][g])
			ws1.write(row, col+12, stdLogVelocity_allFiles[e][f][g])
			ws1.write(row, col+13, aveLogVelocityDiff_allFiles[e][f][g])
			ws1.write(row, col+14, stdLogVelocityDiff_allFiles[e][f][g])
			ws1.write(row, col+15, skewLogVelocityDiff_allFiles[e][f][g])
			ws1.write(row, col+16, kurtLogVelocityDiff_allFiles[e][f][g])
			row += 1 

wkbk_sum.save('Tethering Analysis Summary' + '.xls')

print ('Analysis is done!')